import fitz
import re
import openpyxl
import requests
from tqdm import tqdm  # 进度条
import time

# DEEPSEEK_API_KEY = "sk-26d0fc0bd55045ac9a8f7edcf2cf7718"
# DEEPSEEK_API_URL = "https://api.deepseek.com"
DEEPSEEK_API_KEY = "sk-nwhqbtdjhizqhagulvtgcymarqpxiswaegwjszxyryojktzs"
DEEPSEEK_API_URL = "https://api.siliconflow.cn/v1/chat/completions"


def extract_requirements(text, model="deepseek-ai/DeepSeek-V3"):
    try:
        payload = {
            "model": model,
            "messages": [
                {"role": "system", "content": "你是一个AI助手，专门用于提取与项目相关方需求分析的内容，不做分析和解释"},
                {"role": "user", "content": f"请从以下内容中提取出所有与相关方需求相关的内容，并确保不包含一般性的分析步骤。内容如下：\n\n{text}"}
            ],
            "stream": False,
            "max_tokens": 512,  # 控制文本生成的长度
            "temperature": 0.2,  # 0-1取值，从保守到离谱
            "top_p": 0.7,  # 只考虑概率图中70%的以上的路径
            "top_k": 50,  # 最多挑选五十个词构建概率图
            "frequency_penalty": 0.5,  # 重复机制，0-2，2的惩罚最重，输出的多样性越强
            "n": 1  # 只返回一个输出结果
        }

        headers = {
            "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
            "Content-Type": "application/json"
        }
        print(f"请求内容: {payload}")

        response = requests.post(DEEPSEEK_API_URL, json=payload, headers=headers)

        if response.status_code == 200:
            response_data = response.json()
            extracted_text = response_data['choices'][0]['message']['content'].strip()
            # print(f"gpt的输出: {extracted_text}")

            if "No relevant requirements found" in extracted_text or len(extracted_text) < 10:
                return None
            return extracted_text
        else:
            print(f"请求失败，状态码: {response.status_code}")
            print(response.json())
            return None

    except Exception as e:
        print(f"调用DeepSeek API出错: {e}")
        return None


def extract_chapters_to_excel(pdf_path, output_excel):
    doc = fitz.open(pdf_path)
    text = ""

    for page in doc:
        text += page.get_text("text") + "\n"

    # pattern = re.compile(r"(?m)^\d+(\.\d+)*\s+.+$")  # 分段所有子标题
    pattern = re.compile(r"(?P<title>^\d+\s+.*)", re.MULTILINE)   #只分大章节
    matches = list(pattern.finditer(text))

    chapters = []
    for i in range(len(matches)):
        title = matches[i].group().strip()
        start = matches[i].end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        content = text[start:end].strip()
        full_chapter = f"{title}\n{content}"
        chapters.append([full_chapter])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extracted Chapters"

    for chapter in chapters:
        ws.append(chapter)

    wb.save(output_excel)
    print(f"章节提取完成，已保存到 {output_excel}")

# ========================================================================================================
# def process_excel(input_excel, output_excel):
#     wb = openpyxl.load_workbook(input_excel)
#     ws = wb.active
#
#     for row in tqdm(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1), desc="提取需求中"):
#         cell = row[0]
#         if cell.value:
#             extracted_requirements = extract_requirements(cell.value)
#             # print("提取的需求：",extracted_requirements)
#
#             time.sleep(1)  # 避免API请求过快
#             ws.cell(row=cell.row, column=2, value=extracted_requirements)
#             #
#     wb.save(output_excel)
#     print(f"需求提取完成，已保存到 {output_excel}")
# ========================================================================================================

def process_excel(input_excel, output_excel):
    wb = openpyxl.load_workbook(input_excel)
    ws = wb.active
    error_count = 0

    for row in tqdm(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1), desc="提取需求中"):
        cell = row[0]
        if cell.value:

            print(f"正在处理的内容: {cell.value}")

            # 将内容传给需求提取模型
            extracted_requirements = extract_requirements(cell.value)
            print("llm提取出的需求：",extracted_requirements)

            if extracted_requirements is None:
                error_count += 1
                time.sleep(min(2 ** error_count, 60))  # 如果API忙，可以等待一段时间
            # ========================================================================================================
            else:
                # 计数器，正常情况下不等待
                error_count = 0
                time.sleep(0.5)
            # ========================================================================================================
            # 将提取到的需求填入第二列
            ws.cell(row=cell.row, column=2, value=extracted_requirements)

    wb.save(output_excel)
    print(f"需求提取完成，已保存到 {output_excel}")


pdf_file = "汽车整车信息安全技术要求.pdf"
output_excel = "output_1.xlsx"
extract_chapters_to_excel(pdf_file, output_excel)
process_excel(output_excel, "output_f.xlsx")
