# -*- coding: utf-8 -*-
"""
===================================================================================
Copyright (c) 2025, Yuxuan Zhou & Xiaolong Wang
Author: Yuxuan Zhou & Xiaolong Wang
All rights reserved.
===================================================================================
"""

import openpyxl
import re

class RequirementExtractor:
    def __init__(self, input_excel):
        self.input_excel = input_excel

    def split_titles_and_text(self):
        # 加载 Excel
        wb = openpyxl.load_workbook(self.input_excel)
        ws = wb.active
        new_rows = [["原始段落", "提取需求", "章节号", "需求内容"]]  # 新表格数据，包含表头

        # 章节号匹配正则（假设章节号不是中文）
        chapter_pattern = re.compile(r"^([^\u4e00-\u9fa5]+)(.*)")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):  # 只处理 B 列
            cell = row[0]
            if cell.value:
                text = str(cell.value).strip()

                # 过滤掉 "无明确需求"
                if text == "无明确需求":
                    continue

                match = chapter_pattern.match(text)

                if match:
                    chapter_number = match.group(1).strip()
                    content = match.group(2).strip(": ")
                else:
                    chapter_number = ""
                    content = text  # 如果没有匹配章节号，则直接作为需求内容

                # 获取 A 列（原始段落）
                original_text = ws.cell(row=cell.row, column=1).value

                # 存入新数据
                new_rows.append([original_text, text, chapter_number, content])

        # **清空原表并写入新数据**
        ws.delete_rows(1, ws.max_row)  # 清空原数据
        for row_data in new_rows:
            ws.append(row_data)  # 逐行写入新表格数据

        wb.save(self.input_excel)  # 直接覆盖原文件
        # print(f"✅ 章节号和内容拆分完成，已更新 {self.input_excel}")


if __name__ == "__main__":
    input_excel = "output_2.xlsx"  # 你的 Excel 文件
    extractor = RequirementExtractor(input_excel)
    extractor.split_titles_and_text()
