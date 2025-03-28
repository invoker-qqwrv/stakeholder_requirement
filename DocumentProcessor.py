# -*- coding: utf-8 -*-
"""
===================================================================================
Copyright (c) 2025, Yuxuan Zhou & Xiaolong Wang
Author: Yuxuan Zhou & Xiaolong Wang
All rights reserved.
===================================================================================
"""

import pymupdf as fitz
import re
import openpyxl
import requests
from tqdm import tqdm  # 进度条
import time
from datetime import datetime


class DocumentProcessor:
    def __init__(self, pdf_file, output_excel, api_key, api_url):
        self.pdf_file = pdf_file
        self.output_excel = output_excel
        self.DEEPSEEK_API_KEY = api_key
        self.DEEPSEEK_API_URL = api_url

    def extract_requirements(self, text, model="deepseek-ai/DeepSeek-V3"):
        try:
            payload = {
                "model": model,
                "messages": [
                    {"role": "system",
                     "content": "你现在是一个系统工程师，要从我给你的资料中提取相关方需求。相关方需求（Stakeholder Requirements）**是指项目或系统的利益相关方（如客户、用户、管理层、法规机构等）对其期望、功能、性能、约束等方面的需求，当对某个专业术语进行讲解和定义的时候不可识别为需求。例如：功能需求：系统应支持语音控制。性能需求：响应时间应小于1秒。合规需求：系统必须符合GDPR数据隐私法规。当出现如应该，必须，建议，最好，应满足，必须满足等词的时候必须视为相关方需求。当对需求进行提取时，应结合上下文给输出的需求添加主语，应用条件或相关限制词等，如果在当前输入中没有提取到相关方需求，不需要进行推测，直接输出None"},
                    {"role": "user",
                     "content": f"输出相关方需求的时候连同输入的章节号一起输出。输出时候结合上下文，标题和内容在不要输出到不同的行。如果识别到目录就直接略过，不需要多余的话。内容如下：\n\n{text}"}
                ],
                "stream": False,
                "max_tokens": 1000,  # 控制文本生成的长度
                "temperature": 0.2,  # 0-1取值，从保守到离谱
                "top_p": 0.7,  # 只考虑概率图中70%的以上的路径
                "top_k": 50,  # 最多挑选五十个词构建概率图
                "frequency_penalty": 0.5,  # 重复机制，0-2，2的惩罚最重，输出的多样性越强
                "n": 1  # 只返回一个输出结果
            }

            headers = {
                "Authorization": f"Bearer {self.DEEPSEEK_API_KEY}",
                "Content-Type": "application/json"
            }
            # print(f"请求内容: {payload}")

            response = requests.post(self.DEEPSEEK_API_URL, json=payload, headers=headers)

            if response.status_code == 200:
                response_data = response.json()
                extracted_text = response_data['choices'][0]['message']['content'].strip()

                if "No relevant requirements found" in extracted_text or len(extracted_text) < 10:
                    return None
                return extracted_text
            else:
                print(f"请求失败，状态码: {response.status_code}")
                print(response.json())
                return None

        except Exception as e:
            print(f"调用DeepSeek API出错: {e}")
            return None

    def extract_chapters_to_excel(self):
        doc = fitz.open(self.pdf_file)
        text = ""

        for page in doc:
            text += page.get_text("text") + "\n"

        pattern = re.compile(r"(?P<title>^\d+\s+.*)", re.MULTILINE)  # 只分大章节
        # pattern = re.compile(r"(?m)^\d+(\.\d+)*\s+.+$", re.MULTILINE)  # 分段所有子标题
        matches = list(pattern.finditer(text))

        chapters = []
        for i in range(len(matches)):
            title = matches[i].group().strip()
            start = matches[i].end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
            content = text[start:end].strip()
            full_chapter = f"{title}\n{content}"
            chapters.append([full_chapter])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Chapters"

        for chapter in chapters:
            ws.append(chapter)

        wb.save(self.output_excel)
        # print(f"章节提取完成，已保存到 {self.output_excel}")

    def process_excel(self, input_excel):
        wb = openpyxl.load_workbook(input_excel)
        ws = wb.active
        new_rows = []  # 存储新数据

        for row in tqdm(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1), desc="提取需求中"):
            cell = row[0]
            if cell.value:
                # print(f"正在处理的内容: {cell.value}")

                # 调用 LLM 提取需求
                extracted_requirements = self.extract_requirements(cell.value)
                # print("LLM 提取出的需求：", extracted_requirements)

                # **处理返回值格式**
                if isinstance(extracted_requirements, list):
                    extracted_requirements = "\n".join(extracted_requirements)  # 转换成字符串
                elif extracted_requirements is None or not extracted_requirements.strip():
                    extracted_requirements = "无明确需求"

                time.sleep(0.5)  # 控制请求速率

                # **拆分需求并按行填充**
                requirements_list = extracted_requirements.split("\n")  # 按换行拆分
                first_entry = True  # 标记是否是当前段落的第一行
                for req in requirements_list:
                    if first_entry:
                        new_rows.append([cell.value, req])  # 第一行填充原始段落和需求
                        first_entry = False
                    else:
                        new_rows.append(["", req])  # 后续行只填需求，不重复段落

        # **清空原表格并写入新数据**
        ws.delete_rows(1, ws.max_row)  # 清空原表格
        ws.append(["原始段落", "提取需求"])  # 添加表头
        for row_data in new_rows:
            ws.append(row_data)  # 逐行写入需求

        wb.save(self.output_excel)
        # print(f"需求提取完成，已保存到 {self.output_excel}")


if __name__ == "__main__":
    pdf_file = "汽车整车信息安全技术要求.pdf"
    output_excel = "output_2.xlsx"
    DEEPSEEK_API_KEY = "sk-nwhqbtdjhizqhagulvtgcymarqpxiswaegwjszxyryojktzs"
    DEEPSEEK_API_URL = "https://api.siliconflow.cn/v1/chat/completions"
    processor = DocumentProcessor(pdf_file, output_excel, DEEPSEEK_API_KEY, DEEPSEEK_API_URL)
    processor.extract_chapters_to_excel()
    processor.process_excel(output_excel)
