# -*- coding: utf-8 -*-
"""
===================================================================================
Copyright (c) 2025, Yuxuan Zhou & Xiaolong Wang
Author: Yuxuan Zhou & Xiaolong Wang
All rights reserved.
===================================================================================
"""

import openpyxl
import re
from tqdm import tqdm
from datetime import datetime

class ExcelProcessor:
    def __init__(self, input_excel, output_excel, file_source, file_version, analyst):
        self.input_excel = input_excel
        self.output_excel = output_excel
        self.file_source = file_source
        self.file_version = file_version
        self.analyst = analyst

    def copy_column_to_excel(self):
        try:
            # **读取 input Excel**
            wb_input = openpyxl.load_workbook(self.input_excel, data_only=True)
            ws_input = wb_input.active

            # 读取输出 Excel
            wb_output = openpyxl.load_workbook(self.output_excel)
            sheet_name = "Part1_需求解析表单"

            # 如果没有找到该 sheet，就重新创建一个
            if sheet_name not in wb_output.sheetnames:
                ws_output = wb_output.create_sheet(sheet_name)
            else:
                ws_output = wb_output[sheet_name]

            # **确定插入起始行**
            start_row = 3
            while ws_output.cell(row=start_row, column=4).value:  # 查找第四列的最后一个非空行
                start_row += 1

            # **逐行复制数据**
            for row in tqdm(ws_input.iter_rows(min_row=1, max_row=ws_input.max_row, min_col=3, max_col=4),
                            desc="复制需求到 Part1_需求解析表单"):
                chapter_number = row[0].value  # 第三列
                requirement = row[1].value  # 第四列

                # 获取当前时间
                current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                # 填充数据到输出文件
                if chapter_number and requirement:
                    ws_output.cell(row=start_row, column=1, value=start_row - 2)  # 序号
                    ws_output.cell(row=start_row, column=4, value=chapter_number)  # 章节号
                    ws_output.cell(row=start_row, column=5, value=requirement)  # 需求内容
                    ws_output.cell(row=start_row, column=7, value=current_time)  # 当前时间
                    ws_output.cell(row=start_row, column=2, value=self.file_source)  # 文件来源
                    ws_output.cell(row=start_row, column=3, value=self.file_version)  # 文件版本
                    ws_output.cell(row=start_row, column=6, value=self.analyst)  # 解读分析责任人
                    start_row += 1  # 移动到下一行

            # **保存 Excel 文件**
            wb_output.save(self.output_excel)
            wb_output.close()
            wb_input.close()

            # print(f"✅ 数据复制并保存完成！")

        except PermissionError:
            print("❌ 错误：Excel 文件被占用，请关闭 Excel 后重试！")
        except Exception as e:
            print(f"❌ 发生错误：{e}")

if __name__ == "__main__":
    input_excel = "output_2.xlsx"
    output_excel = "顾客要求跟踪矩阵_1.xlsx"

    #文件名，版本号，解析人姓名天灾此处
    file_source = "《汽车整车信息安全技术要求》"
    file_version = "V1.0"
    analyst = "马牛逼"

    processor = ExcelProcessor(input_excel, output_excel, file_source, file_version, analyst)
    processor.copy_column_to_excel()
